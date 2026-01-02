# Copyright (c) 2019, Frappe Technologies and contributors
# For license information, please see license.txt


import csv
import io
import json
import re
from datetime import date, datetime

import frappe
import openpyxl
from frappe import _
from frappe.core.doctype.data_import.data_import import DataImport
from frappe.core.doctype.data_import.importer import Importer, ImportFile
from frappe.utils.background_jobs import enqueue
from frappe.utils.file_manager import get_file, save_file
from frappe.utils.xlsxutils import ILLEGAL_CHARACTERS_RE, handle_html
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

INVALID_VALUES = ("", None)


class MoldovaBankStatementImport(DataImport):
	# begin: auto-generated types
	# This code is auto-generated. Do not modify anything in this block.

	from typing import TYPE_CHECKING

	if TYPE_CHECKING:
		from frappe.types import DF

		bank: DF.Link | None
		bank_account: DF.Link
		company: DF.Link
		custom_delimiters: DF.Check
		delimiter_options: DF.Data | None
		google_sheets_url: DF.Data | None
		import_file: DF.Attach | None
		import_dbo_fromat: DF.Check
		import_type: DF.Literal["", "Insert New Records", "Update Existing Records"]
		mute_emails: DF.Check
		reference_doctype: DF.Link
		show_failed_logs: DF.Check
		status: DF.Literal["Pending", "Success", "Partial Success", "Error"]
		submit_after_import: DF.Check
		template_options: DF.Code | None
		template_warnings: DF.Code | None
	# end: auto-generated types

	def __init__(self, *args, **kwargs):
		super().__init__(*args, **kwargs)

	def validate(self):
		doc_before_save = self.get_doc_before_save()
		if (
			not (self.import_file or self.google_sheets_url)
			or (doc_before_save and doc_before_save.import_file != self.import_file)
			or (doc_before_save and doc_before_save.google_sheets_url != self.google_sheets_url)
		):
			template_options_dict = {}
			column_to_field_map = {}
			bank = frappe.get_doc("Bank", self.bank)
			for i in bank.bank_transaction_mapping:
				column_to_field_map[i.file_field] = i.bank_transaction_field
			template_options_dict["column_to_field_map"] = column_to_field_map
			self.template_options = json.dumps(template_options_dict)

			self.template_warnings = ""

		if self.import_file and not self.import_file.lower().endswith(".txt"):
			self.validate_import_file()
			self.validate_google_sheets_url()

	def start_import(self):
		preview = frappe.get_doc("Moldova Bank Statement Import", self.name).get_preview_from_template(
			self.import_file, self.google_sheets_url
		)

		if "Bank Account" not in json.dumps(preview["columns"]):
			frappe.throw(_("Please add the Bank Account column"))

		from frappe.utils.background_jobs import is_job_enqueued
		from frappe.utils.scheduler import is_scheduler_inactive

		if is_scheduler_inactive() and not frappe.flags.in_test:
			frappe.throw(_("Scheduler is inactive. Cannot import data."), title=_("Scheduler Inactive"))

		job_id = f"moldova_bank_statement_import::{self.name}"
		if not is_job_enqueued(job_id):
			enqueue(
				start_import,
				queue="default",
				timeout=6000,
				event="data_import",
				job_id=job_id,
				data_import=self.name,
				bank_account=self.bank_account,
				import_file_path=self.import_file,
				google_sheets_url=self.google_sheets_url,
				bank=self.bank,
				template_options=self.template_options,
				now=frappe.conf.developer_mode or frappe.flags.in_test,
			)
			return True

		return False

@frappe.whitelist()
def convert_dbo_to_csv(data_import, dbo_file_path):
	from frappe.utils import cstr

	doc = frappe.get_doc("Moldova Bank Statement Import", data_import)

	_file_doc, content = get_file(dbo_file_path)

	is_dbo = is_dbo_format(content)
	if not is_dbo:
		frappe.throw(_("The uploaded file does not appear to be in valid DBO format."))

	if is_dbo and not doc.import_dbo_fromat:
		frappe.throw(_("DBO file detected. Please enable 'Import DBO Format' to proceed."))

	try:
		transactions = parse_dbo(content)
	except Exception as e:
		frappe.throw(_("Failed to parse DBO format. Error: {0}").format(str(e)))

	if not transactions:
		frappe.throw(_("Parsed file is not in valid DBO format or contains no transactions."))

	# Use in-memory file buffer instead of writing to temp file
	csv_buffer = io.StringIO()
	writer = csv.writer(csv_buffer)

	headers = [
		"Date",
		"Deposit",
		"Withdrawal",
		"Description",
		"Reference Number",
		"Bank Account",
		"Currency",
		"Party Type",
		"Party",
		"Party Name/Account Holder (Bank Statement)",
		"Party Account No. (Bank Statement)",
		"Party IBAN (Bank Statement)",
	]

	writer.writerow(headers)

	for txn in transactions:

		# Date & currency safe formatting
		dt = txn.get("date")
		date_str = cstr(dt) if dt else ""

		deposit = txn.get("deposit") or 0
		withdrawal = txn.get("withdrawal") or 0
		description = (txn.get("description") or "").replace("\r\n", "\n") or ""
		reference = txn.get("reference_number") or ""
		currency = txn.get("currency", "")
		
        # Party resolution based on IDNO
		party_type, party = resolve_party_by_idno(txn)

		# Counterparty account split into account number vs IBAN
		cp_account_raw = (txn.get("cp_account") or "").strip()
		is_iban = is_iban_valid(cp_account_raw.replace(" ", ""))

		party_iban = cp_account_raw if is_iban else ""
		party_account_no = "" if is_iban else cp_account_raw

		writer.writerow([
			date_str, 
			deposit, 
			withdrawal, 
			description, 
			reference, 
			doc.bank_account, 
			currency,
            party_type or "",
            party or "",
            (txn.get("cp_name") or ""),
            party_account_no,
            party_iban,
        ])

	# Prepare in-memory CSV for upload
	csv_content = csv_buffer.getvalue().encode("utf-8")
	csv_buffer.close()

	filename = f"{frappe.utils.now_datetime().strftime('%Y%m%d%H%M%S')}_converted_dbo.csv"

	# Save to File Manager
	saved_file = save_file(filename, csv_content, doc.doctype, doc.name, is_private=True, df="import_file")

	return saved_file.file_url


@frappe.whitelist()
def get_preview_from_template(data_import, import_file=None, google_sheets_url=None):
	return frappe.get_doc("Moldova Bank Statement Import", data_import).get_preview_from_template(
		import_file, google_sheets_url
	)


@frappe.whitelist()
def form_start_import(data_import):
	job_id = frappe.get_doc("Moldova Bank Statement Import", data_import).start_import()
	return job_id is not None


@frappe.whitelist()
def download_errored_template(data_import_name):
	data_import = frappe.get_doc("Moldova Bank Statement Import", data_import_name)
	data_import.export_errored_rows()


@frappe.whitelist()
def download_import_log(data_import_name):
	return frappe.get_doc("Moldova Bank Statement Import", data_import_name).download_import_log()


def is_dbo_format(content: str) -> bool:
	"""Check if the content has key DBO tags"""
	required_tags = ["DocStart", "DocEnd", "BEGINDATE", "ENDDATE"]
	return all(tag in content for tag in required_tags)


def parse_data_from_template(raw_data):
	data = []

	for _i, row in enumerate(raw_data):
		if all(v in INVALID_VALUES for v in row):
			# empty row
			continue

		data.append(row)

	return data


def start_import(data_import, bank_account, import_file_path, google_sheets_url, bank, template_options):
	"""This method runs in background job"""

	update_mapping_db(bank, template_options)

	data_import = frappe.get_doc("Moldova Bank Statement Import", data_import)
	file = import_file_path if import_file_path else google_sheets_url

	import_file = ImportFile("Bank Transaction", file=file, import_type="Insert New Records")

	data = parse_data_from_template(import_file.raw_data)
	# Importer expects 'Data Import' class, which has 'payload_count' attribute
	if not data_import.get("payload_count"):
		data_import.payload_count = len(data) - 1

	if import_file_path:
		add_bank_account(data, bank_account)
		write_files(import_file, data)

	try:
		i = Importer(data_import.reference_doctype, data_import=data_import)
		i.import_data()
	except Exception:
		frappe.db.rollback()
		data_import.db_set("status", "Error")
		data_import.log_error("Moldova Bank Statement Import failed")
	finally:
		frappe.flags.in_import = False

	frappe.publish_realtime("data_import_refresh", {"data_import": data_import.name})


def update_mapping_db(bank, template_options):
	"""Update bank transaction mapping database with template options."""
	bank = frappe.get_doc("Bank", bank)
	for d in bank.bank_transaction_mapping:
		d.delete()

	for d in json.loads(template_options)["column_to_field_map"].items():
		bank.append("bank_transaction_mapping", {"bank_transaction_field": d[1], "file_field": d[0]})

	bank.save()


def add_bank_account(data, bank_account):
	"""Add bank account information to data rows."""
	bank_account_loc = None
	if "Bank Account" not in data[0]:
		data[0].append("Bank Account")
	else:
		for loc, header in enumerate(data[0]):
			if header == "Bank Account":
				bank_account_loc = loc

	for row in data[1:]:
		if bank_account_loc:
			row[bank_account_loc] = bank_account
		else:
			row.append(bank_account)


def write_files(import_file, data):
	"""Write processed data to CSV or Excel files."""
	full_file_path = import_file.file_doc.get_full_path()
	parts = import_file.file_doc.get_extension()
	extension = parts[1]
	extension = extension.lstrip(".")

	if extension == "csv":
		with open(full_file_path, "w", newline="") as file:
			writer = csv.writer(file)
			writer.writerows(data)
	elif extension in ("xlsx", "xls"):
		write_xlsx(data, "trans", file_path=full_file_path)


def write_xlsx(data, sheet_name, wb=None, column_widths=None, file_path=None):
	"""Write data to Excel file with formatting."""
	# from xlsx utils with changes
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=True)

	ws = wb.create_sheet(sheet_name, 0)

	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	row1 = ws.row_dimensions[1]
	row1.font = Font(name="Calibri", bold=True)

	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)

			clean_row.append(value)

		ws.append(clean_row)

	wb.save(file_path)
	return True


@frappe.whitelist()
def get_import_status(docname):
	import_status = {}

	data_import = frappe.get_doc("Moldova Bank Statement Import", docname)
	import_status["status"] = data_import.status

	logs = frappe.get_all(
		"Data Import Log",
		fields=["count(*) as count", "success"],
		filters={"data_import": docname},
		group_by="success",
	)

	total_payload_count = 0

	for log in logs:
		total_payload_count += log.get("count", 0)
		if log.get("success"):
			import_status["success"] = log.get("count")
		else:
			import_status["failed"] = log.get("count")

	import_status["total_records"] = total_payload_count

	return import_status


@frappe.whitelist()
def get_import_logs(docname: str):
	frappe.has_permission("Moldova Bank Statement Import", throw=True)

	return frappe.get_all(
		"Data Import Log",
		fields=["success", "docname", "messages", "exception", "row_indexes"],
		filters={"data_import": docname},
		limit_page_length=5000,
		order_by="log_index",
	)


@frappe.whitelist()
def upload_bank_statement(**args):
	args = frappe._dict(args)
	bsi = frappe.new_doc("Moldova Bank Statement Import")

	if args.company:
		bsi.update(
			{
				"company": args.company,
			}
		)

	if args.bank_account:
		bsi.update({"bank_account": args.bank_account})

	return bsi

def parse_date(value: str):
    """Parse date in DD.MM.YYYY format into python date."""
	
    from frappe.utils import getdate

    value = (value or "").strip()
    if not value:
        return None
    return getdate(value)

def parse_dbo(content: str):
    """Parse DBO formatted bank statement content into transactions."""
    # This is a placeholder implementation. The actual parsing logic will depend on the DBO format specification.
    # For demonstration, let's assume we have a simple parser that extracts transactions based on known tags.

    from frappe.utils import flt

    lines = [ln.strip() for ln in content.splitlines() if ln.strip()]

    header = {}
    docs = []
    current_doc = None
    in_account_section = False

    for line in lines:
        if line == "SECTIONACCOUNTSTART":
            in_account_section = True
            continue

        if line == "SECTIONACCOUNTSTOP":
            in_account_section = False
            continue

        if line == "DocStart":
            current_doc = {}
            continue

        if line == "DocEnd":
            if current_doc:
                docs.append(current_doc)
                current_doc = None
            continue

        if "=" in line:
            key, value = line.split("=", 1)
            key = key.strip().upper()
            value = value.strip()

            if in_account_section:
                # Account-level fields (ACCOUNT, STARTREST, STOPREST, CURRCODE, etc.)
                header[key] = value

            elif current_doc is not None:
                # Document-level fields (DOCUMENTNUMBER, AMOUNT, GROUND, etc.)
                current_doc[key] = value

            else:
                # Global header fields (BEGINDATE, ENDDATE)
                header[key] = value

    """
    Convert parsed document blocks into a list of transaction dicts ready
    to be converted into Bank Transaction documents.

    Expected transaction dict keys:
    - date
    - description
    - deposit
    - withdrawal
    - bank_balance
    - reference_number
    - currency
    - cp_role              (Payer / Receiver)
    - cp_name
    - cp_account
    - cp_idno
    - cp_bank
    - cp_bank_bic
    """
    account_iban = header.get("ACCOUNT")
    opening_balance = flt(header.get("STARTREST") or 0)
    closing_balance_bank = flt(header.get("STOPREST") or 0)
    begin_date_str = header.get("BEGINDATE")
    end_date_str = header.get("ENDDATE")
    currency_code = (header.get("CURRCODE") or "").strip() or None

    transactions: list[dict] = []
    running_balance = opening_balance
    from_date = None
    to_date = None

    for doc in docs:
        # Raw fields from statement
        document_number = (doc.get("DOCUMENTNUMBER") or "").strip()
        document_date_str = (doc.get("DOCUMENTDATE") or "").strip()
        date_written_str = (doc.get("DATEWRITTEN") or "").strip()

        posting_date = parse_date(document_date_str or date_written_str)

        amount = flt(doc.get("AMOUNT") or 0)

        payer_account = (doc.get("PAYERACCOUNT") or "").strip()
        receiver_account = (doc.get("RECEIVERACCOUNT") or "").strip()

        payer_name = (doc.get("PAYER") or "").strip()
        receiver_name = (doc.get("RECEIVER") or "").strip()

        payer_fcode = (doc.get("PAYERFCODE") or "").strip()
        receiver_fcode = (doc.get("RECEIVERFCODE") or "").strip()

        payer_bank = (doc.get("PAYERBANK") or "").strip()
        receiver_bank = (doc.get("RECEIVERBANK") or "").strip()

        payer_bank_bic = (doc.get("PAYERBANKBIC") or "").strip()
        receiver_bank_bic = (doc.get("RECEIVERBANKBIC") or "").strip()

        oper_type = (doc.get("OPERTYPE") or "").strip()
        transaction_code = (doc.get("TRANSACTIONCODE") or "").strip()

        base_ground = (doc.get("GROUND") or "").strip()

        deposit = 0.0
        withdrawal = 0.0
        cp_role = ""
        cp_name = ""
        cp_account = ""
        cp_idno = ""
        cp_bank = ""
        cp_bank_bic = ""

        # Direction: if our account is payer -> withdrawal, if receiver -> deposit
        if account_iban and payer_account == account_iban and amount:
            withdrawal = amount
            running_balance -= amount
            cp_role = "Receiver"
            cp_name = receiver_name
            cp_account = receiver_account
            cp_idno = receiver_fcode
            cp_bank = receiver_bank
            cp_bank_bic = receiver_bank_bic

        elif account_iban and receiver_account == account_iban and amount:
            deposit = amount
            running_balance += amount
            cp_role = "Payer"
            cp_name = payer_name
            cp_account = payer_account
            cp_idno = payer_fcode
            cp_bank = payer_bank
            cp_bank_bic = payer_bank_bic

        # Track min/max posting date
        if posting_date:
            if not from_date or posting_date < from_date:
                from_date = posting_date
            if not to_date or posting_date > to_date:
                to_date = posting_date

        # Build description in required format:
        # 1) GROUND
        # 2) empty line
        # 3) "Amount: ..."
        # 4) "Document Number: ..."
        # 5) "Date Written: ..."
        # 6) "<Receiver/Payer>: ..."
        # 7) "<Receiver/Payer> IDNO: ..."
        # 8) "<Receiver/Payer> Account: ..."
        # 9) "<Receiver/Payer> Bank: ..."
        # 10) "<Receiver/Payer> Bank BIC: ..."
        desc_lines = []

        if base_ground:
            desc_lines.append(base_ground)

        # Empty line
        if desc_lines:
            desc_lines.append("")

        # Amount
        if amount:
            desc_lines.append(f"Amount: {amount:.2f}")

        # Document number
        if document_number:
            desc_lines.append(f"Document Number: {document_number}")

        # Date written (string as in statement)
        if date_written_str:
            desc_lines.append(f"Date Written: {date_written_str}")

        # Counterparty block
        if cp_role and (cp_name or cp_account or cp_idno or cp_bank or cp_bank_bic):
            # Name
            if cp_name:
                desc_lines.append(f"{cp_role}: {cp_name}")
            # IDNO
            if cp_idno:
                desc_lines.append(f"{cp_role} IDNO: {cp_idno}")
            # Account
            if cp_account:
                desc_lines.append(f"{cp_role} Account: {cp_account}")
            # Bank
            if cp_bank:
                desc_lines.append(f"{cp_role} Bank: {cp_bank}")
            # Bank BIC
            if cp_bank_bic:
                desc_lines.append(f"{cp_role} Bank BIC: {cp_bank_bic}")

        # Optional technical info
        if oper_type or transaction_code:
            tech_parts = []
            if oper_type:
                tech_parts.append(f"OpType: {oper_type}")
            if transaction_code:
                tech_parts.append(f"TxnCode: {transaction_code}")
            if tech_parts:
                desc_lines.append(" / ".join(tech_parts))

        description = "\n".join(desc_lines)

        transactions.append({
            "date": posting_date,
            "description": description,
            "deposit": deposit,
            "withdrawal": withdrawal,
            "bank_balance": running_balance,
            "reference_number": document_number,
            "currency": currency_code,
            "cp_role": cp_role,
            "cp_name": cp_name,
            "cp_account": cp_account,
            "cp_idno": cp_idno,
            "cp_bank": cp_bank,
            "cp_bank_bic": cp_bank_bic,
        })

    # Prefer dates from header if available, otherwise derived from documents
    #statement_from_date = parse_date(begin_date_str) or from_date
    #statement_to_date = parse_date(end_date_str) or to_date

    #opening_balance = opening_balance
    # If bank reports closing balance, use it; otherwise trust running balance
    #closing_balance = closing_balance_bank or running_balance

    return transactions

def resolve_party_by_idno(tx: dict) -> tuple[str, str]:
	"""Resolve party_type and party name used in CSV based on IDNO.

	This follows the same rules as ``_assign_party_by_idno`` but instead of
	mutating a Bank Transaction document it just returns the match so that
	we can include it when generating the CSV.
	"""
	cp_idno = (tx.get("cp_idno") or "").strip()
	if not cp_idno:
		return "", ""

	try:
		settings = frappe.get_single("Moldova Banking Settings")
	except Exception:
		return "", ""

	customer_idno_field = (getattr(settings, "customer_idno_field", None) or "").strip()
	supplier_idno_field = (getattr(settings, "supplier_idno_field", None) or "").strip()

	# Incoming payment -> Customer
	if tx.get("deposit") and not tx.get("withdrawal") and customer_idno_field:
		customer = frappe.get_all(
			"Customer",
			filters={customer_idno_field: cp_idno},
			pluck="name",
			limit=1,
		)
		if customer:
			return "Customer", customer[0]

	# Outgoing payment -> Supplier
	if tx.get("withdrawal") and not tx.get("deposit") and supplier_idno_field:
		supplier = frappe.get_all(
			"Supplier",
			filters={supplier_idno_field: cp_idno},
			pluck="name",
			limit=1,
		)
		if supplier:
			return "Supplier", supplier[0]

	return "", ""

def is_iban_valid(iban_string):
    """
    Checks if a string is a valid IBAN format using a regular expression.
    Returns True if valid, False otherwise.
    """
    # Remove spaces and convert to uppercase for consistent matching
    cleaned_iban = iban_string.replace(" ", "").upper()

    # A general IBAN regex pattern (adjust for more specific country rules if needed)
    iban_pattern = r"^[A-Z]{2}\d{2}[A-Z0-9]{11,30}$"

    if re.fullmatch(iban_pattern, cleaned_iban):
        return True
    else:
        return False